from fastapi import FastAPI, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from typing import List
import openpyxl
import io


import models, database, schemas, crud, google_sheets
from database import engine, get_db

# Create database tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="U-Planner API")

# --- Google Sheets Sync ---
@app.post("/sync/google-sheets/")
def sync_google_sheets(db: Session = Depends(get_db)):
    service = google_sheets.GoogleSheetsService()
    return service.sync_full_data(db)

@app.get("/")
def read_root():
    return {"message": "Welcome to U-Planner API"}

# --- Teachers ---
@app.get("/teachers/", response_model=List[schemas.Teacher])
def read_teachers(db: Session = Depends(get_db)):
    return db.query(models.Teacher).all()

@app.post("/teachers/", response_model=schemas.Teacher)
def create_teacher(teacher: schemas.TeacherBase, db: Session = Depends(get_db)):
    db_teacher = models.Teacher(full_name=teacher.full_name, rut=teacher.rut)
    db.add(db_teacher)
    try:
        db.commit()
        db.refresh(db_teacher)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")
    return db_teacher

@app.put("/teachers/{teacher_id}", response_model=schemas.Teacher)
def update_teacher(teacher_id: int, teacher: schemas.TeacherBase, db: Session = Depends(get_db)):
    db_teacher = db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    if not db_teacher:
        raise HTTPException(status_code=404, detail="Docente no encontrado")
    db_teacher.full_name = teacher.full_name
    db_teacher.rut = teacher.rut
    try:
        db.commit()
        db.refresh(db_teacher)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")
    return db_teacher

@app.post("/teachers/upload-excel/")
async def upload_teachers_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    created = 0
    skipped = 0
    errors = []
    seen_ruts = set()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        
        full_name = str(row[0]).strip() if row[0] else None
        rut = str(row[1]).strip() if row[1] else None
        
        if not full_name or not rut or full_name == "None" or rut == "None":
            continue
        
        # Skip duplicates within the same file
        if rut in seen_ruts:
            skipped += 1
            continue
        seen_ruts.add(rut)
        
        # Check if RUT already exists in DB
        existing = db.query(models.Teacher).filter(models.Teacher.rut == rut).first()
        if existing:
            skipped += 1
            continue
        
        teacher = models.Teacher(full_name=full_name, rut=rut)
        db.add(teacher)
        created += 1
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar en la base de datos: {str(e)}")
    
    return {
        "message": "Importación completada",
        "created": created,
        "skipped": skipped,
        "errors": errors
    }

# --- Schedules ---
@app.post("/schedules/", response_model=schemas.Schedule)
def create_schedule(schedule: schemas.ScheduleBase, db: Session = Depends(get_db)):
    # validation happens inside crud.create_schedule
    return crud.create_schedule(db=db, schedule=schedule)

@app.get("/schedules/", response_model=List[schemas.Schedule])
def read_schedules(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.get_schedules(db=db, skip=skip, limit=limit)

# --- Subjects ---
@app.post("/subjects/", response_model=schemas.Subject)
def create_subject(subject: schemas.SubjectBase, db: Session = Depends(get_db)):
    db_subject = models.Subject(**subject.model_dump())
    db.add(db_subject)
    db.commit()
    db.refresh(db_subject)
    return db_subject

@app.get("/subjects/", response_model=List[schemas.Subject])
def read_subjects(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return db.query(models.Subject).offset(skip).limit(limit).all()

# --- Rooms ---
@app.post("/rooms/", response_model=schemas.Room)
def create_room(room: schemas.RoomBase, db: Session = Depends(get_db)):
    db_room = models.Room(**room.model_dump())
    db.add(db_room)
    db.commit()
    db.refresh(db_room)
    return db_room

@app.get("/rooms/", response_model=List[schemas.Room])
def read_rooms(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return db.query(models.Room).offset(skip).limit(limit).all()

@app.get("/days/", response_model=List[schemas.Day])
def read_days(db: Session = Depends(get_db)):
    return db.query(models.Day).all()

@app.get("/time-modules/", response_model=List[schemas.TimeModule])
def read_time_modules(db: Session = Depends(get_db)):
    return db.query(models.TimeModule).all()

@app.get("/academic-schedules/", response_model=List[schemas.AcademicSchedule])
def read_academic_schedules(skip: int = 0, limit: int = 5000, db: Session = Depends(get_db)):
    return db.query(models.AcademicSchedule).offset(skip).limit(limit).all()





@app.put("/rooms/{room_id}", response_model=schemas.Room)
def update_room(room_id: int, room: schemas.RoomBase, db: Session = Depends(get_db)):
    db_room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not db_room:
        raise HTTPException(status_code=404, detail="Sala no encontrada")
    db_room.code = room.code
    db_room.name = room.name
    db_room.capacity = room.capacity
    try:
        db.commit()
        db.refresh(db_room)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")
    return db_room

@app.post("/rooms/upload-excel/")
async def upload_rooms_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    created = 0
    skipped = 0
    seen_codes = set()
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 3:
            continue
        
        code = str(row[0]).strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        cap_raw = row[2]
        
        if not code or not name or code == "None" or name == "None":
            continue
        
        try:
            capacity = int(cap_raw)
        except (TypeError, ValueError):
            continue
        
        if code in seen_codes:
            skipped += 1
            continue
        seen_codes.add(code)
        
        existing = db.query(models.Room).filter(models.Room.code == code).first()
        if existing:
            skipped += 1
            continue
        
        room = models.Room(code=code, name=name, capacity=capacity)
        db.add(room)
        created += 1
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar: {str(e)}")
    
    return {
        "message": "Importación completada",
        "created": created,
        "skipped": skipped
    }

# --- Room Types ---
@app.get("/room-types/", response_model=List[schemas.RoomType])
def read_room_types(db: Session = Depends(get_db)):
    return db.query(models.RoomType).all()

# --- Faculties & Reports ---
@app.get("/faculties/", response_model=List[schemas.Faculty])
def read_faculties(db: Session = Depends(get_db)):
    return crud.get_faculties(db)

@app.get("/reports/schedules/faculty/{faculty_id}", response_model=List[schemas.Schedule])
def get_faculty_report(faculty_id: int, db: Session = Depends(get_db)):
    return crud.get_schedules_by_faculty(db, faculty_id)

@app.get("/reports/export/faculty/{faculty_id}")
def export_faculty_report(faculty_id: int, db: Session = Depends(get_db)):
    import csv
    import io
    from fastapi.responses import StreamingResponse
    
    schedules = crud.get_schedules_by_faculty(db, faculty_id)
    faculty = db.query(models.Faculty).filter(models.Faculty.id == faculty_id).first()
    faculty_name = faculty.name if faculty else "report"
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Asignatura", "Docente", "Sala", "Bloque"])
    
    for s in schedules:
        writer.writerow([
            s.id,
            s.subject.name,
            s.teacher.user.username,
            s.room.name,
            f"{s.time_block.day_of_week} ({s.time_block.start_time}-{s.time_block.end_time})"
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=horario_{faculty_name}.csv"}
    )
